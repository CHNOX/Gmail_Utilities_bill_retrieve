"""
Gmail Utilities Bill Retrieve
==============================
Legge la configurazione da settings.json e per ogni mittente:
  - recupera tutte le email
  - estrae i valori delle chiavi configurate dal corpo email e/o dagli allegati PDF

Genera un file Excel con una colonna per ogni chiave di estrazione.

PREREQUISITI:
1. Python 3.7+
2. pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib openpyxl pdfplumber

3. Scarica credentials.json da Google Cloud Console (vedi README).

4. Configura settings.json con i mittenti, le chiavi da estrarre e la sorgente
   ("body", "pdf", "both").
"""

import base64
import io
import json
import os
import re
import sys
from datetime import datetime
from html.parser import HTMLParser

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

CREDENTIALS_FILE = "credentials.json"
TOKEN_FILE        = "token.json"
SETTINGS_FILE     = "settings.json"
BATCH_SIZE        = 500


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

def _parse_date_setting(value, field_name):
    """Valida e converte una data in formato YYYY-MM-DD; esce se il formato è errato."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        print(f"ERRORE: '{field_name}' in date_range deve essere nel formato YYYY-MM-DD (es. 2024-01-01).")
        sys.exit(1)


def load_settings():
    """
    Carica settings.json.
    Restituisce (senders_config, date_range) dove date_range è un dict
    {"from": date|None, "to": date|None}.
    """
    if not os.path.exists(SETTINGS_FILE):
        print(f"ERRORE: File '{SETTINGS_FILE}' non trovato.")
        print("Crea il file con la lista dei mittenti e delle chiavi da estrarre.")
        sys.exit(1)

    with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
        settings = json.load(f)

    senders = settings.get("senders", [])
    if not senders:
        print(f"ERRORE: Nessun mittente configurato in '{SETTINGS_FILE}'.")
        sys.exit(1)

    # Normalizza le chiavi in uppercase (case-insensitive, colonna Excel condivisa)
    for s in senders:
        s["extract_bill"] = [k.upper() for k in s.get("extract_bill", [])]
        if s.get("extract_customer"):
            s["extract_customer"] = s["extract_customer"].upper()
        if s.get("extract_period"):
            s["extract_period"] = s["extract_period"].upper()

    # Legge il range di date globale (entrambi i campi opzionali)
    dr_raw     = settings.get("date_range", {})
    date_range = {
        "from": _parse_date_setting(dr_raw.get("from"), "from"),
        "to":   _parse_date_setting(dr_raw.get("to"),   "to"),
    }

    return senders, date_range


def collect_all_keys(senders_config):
    """Restituisce la lista ordinata di tutte le chiavi uniche tra i mittenti."""
    seen = {}
    for s in senders_config:
        for key in s.get("extract_bill", []):
            seen[key] = None  # dict per preservare l'ordine di inserimento
    return list(seen)


# ---------------------------------------------------------------------------
# Autenticazione
# ---------------------------------------------------------------------------

def authenticate():
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("Token scaduto, rinnovo in corso...")
            creds.refresh(Request())
        else:
            if not os.path.exists(CREDENTIALS_FILE):
                print(f"ERRORE: File '{CREDENTIALS_FILE}' non trovato!")
                sys.exit(1)
            print("Apertura browser per autorizzazione...")
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
        print("Autenticazione completata e salvata.")

    return creds


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_date(date_str):
    if not date_str:
        return None
    date_str = re.sub(r'\s*\([^)]*\)\s*$', '', date_str.strip())
    for fmt in (
        "%a, %d %b %Y %H:%M:%S %z",
        "%a, %d %b %Y %H:%M:%S",
        "%d %b %Y %H:%M:%S %z",
        "%d %b %Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


class _HTMLStripper(HTMLParser):
    # Tag che nella resa visiva iniziano una nuova riga — aggiungere \n garantisce
    # che il testo estratto abbia separatori semantici utili per text_only extraction.
    _NEWLINE_TAGS = {"br", "p", "div", "tr", "li", "td", "th",
                     "h1", "h2", "h3", "h4", "h5", "h6",
                     "section", "article", "header", "footer"}

    def __init__(self):
        super().__init__()
        self._parts = []

    def handle_starttag(self, tag, _attrs):
        if tag.lower() in self._NEWLINE_TAGS:
            self._parts.append("\n")

    def handle_endtag(self, tag):
        if tag.lower() in self._NEWLINE_TAGS:
            self._parts.append("\n")

    def handle_data(self, data):
        self._parts.append(data)

    def get_text(self):
        return " ".join(self._parts)


def _html_to_text(html_content):
    s = _HTMLStripper()
    s.feed(html_content)
    return s.get_text()


class _LinkExtractor(HTMLParser):
    """Estrae tutti i link (testo visibile, href) dai tag <a> in HTML."""
    def __init__(self):
        super().__init__()
        self._links      = []
        self._cur_href   = None
        self._cur_text   = []

    def handle_starttag(self, tag, attrs):
        if tag == "a":
            href = dict(attrs).get("href", "")
            if href and href.startswith("http"):
                self._cur_href = href
                self._cur_text = []

    def handle_data(self, data):
        if self._cur_href is not None:
            self._cur_text.append(data)

    def handle_endtag(self, tag):
        if tag == "a" and self._cur_href:
            self._links.append((" ".join(self._cur_text).strip(), self._cur_href))
            self._cur_href = None

    def get_links(self):
        return self._links


def _extract_html_body(payload):
    """Restituisce il corpo HTML grezzo del messaggio (non stripped)."""
    mime = payload.get("mimeType", "")
    if mime == "text/html":
        raw = payload.get("body", {}).get("data", "")
        if raw:
            return base64.urlsafe_b64decode(raw + "==").decode("utf-8", errors="replace")
    for part in payload.get("parts", []):
        pt  = part.get("mimeType", "")
        raw = part.get("body", {}).get("data", "")
        if pt == "text/html" and raw:
            return base64.urlsafe_b64decode(raw + "==").decode("utf-8", errors="replace")
        if part.get("parts"):
            nested = _extract_html_body(part)
            if nested:
                return nested
    return ""


def _find_links(payload):
    """Restituisce lista di (testo, url) dai link <a> nel corpo HTML."""
    extractor = _LinkExtractor()
    extractor.feed(_extract_html_body(payload))
    return extractor.get_links()


def _select_link(links, link_text=None):
    """
    Sceglie l'URL migliore dalla lista.
    Se link_text è fornito, cerca il primo link il cui testo o URL lo contiene
    (case-insensitive). Altrimenti restituisce il primo link disponibile.
    """
    if not links:
        return None
    if link_text:
        lt = link_text.lower()
        for text, url in links:
            if lt in text.lower() or lt in url.lower():
                return url
    return links[0][1]


def _download_url_as_text(url):
    """
    Scarica l'URL e restituisce il testo estratto.
    Gestisce PDF (pdfplumber) e HTML (stripping tag).
    """
    import requests
    try:
        resp = requests.get(url, timeout=30, allow_redirects=True)
        resp.raise_for_status()
        ctype = resp.headers.get("Content-Type", "").lower()
        if "pdf" in ctype or url.lower().split("?")[0].endswith(".pdf"):
            return _pdf_bytes_to_text(resp.content)
        if "html" in ctype:
            return _html_to_text(resp.text)
        return resp.text
    except Exception as e:
        print(f"    Errore download link ({url[:70]}): {e}")
        return ""


def extract_body_text(payload):
    """Estrae ricorsivamente il testo dal payload Gmail (plain > html > multipart)."""
    mime = payload.get("mimeType", "")

    if mime == "text/plain":
        raw = payload.get("body", {}).get("data", "")
        if raw:
            return base64.urlsafe_b64decode(raw + "==").decode("utf-8", errors="replace")

    if mime == "text/html":
        raw = payload.get("body", {}).get("data", "")
        if raw:
            return _html_to_text(
                base64.urlsafe_b64decode(raw + "==").decode("utf-8", errors="replace")
            )

    plain, html = "", ""
    for part in payload.get("parts", []):
        pt = part.get("mimeType", "")
        raw = part.get("body", {}).get("data", "")
        if pt == "text/plain" and raw:
            plain = base64.urlsafe_b64decode(raw + "==").decode("utf-8", errors="replace")
        elif pt == "text/html" and raw:
            html = _html_to_text(
                base64.urlsafe_b64decode(raw + "==").decode("utf-8", errors="replace")
            )
        elif part.get("parts"):
            nested = extract_body_text(part)
            plain = plain or nested

    return plain or html


def _parse_amount(value):
    """
    Converte un valore estratto (es. '€ 1.234,56' o '1234.56') in float.
    Restituisce None se il valore non è numerico.
    """
    if not value:
        return None
    cleaned = re.sub(r"[€\s]", "", value)
    # formato italiano: 1.234,56 → separatore migliaia '.' e decimale ','
    if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", cleaned):
        cleaned = cleaned.replace(".", "").replace(",", ".")
    else:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def extract_value_for_key(text, key, text_only=False):
    """
    Cerca `key` nel testo e restituisce il valore associato.
    Strategia (skippata se text_only=True):
      1. Importo con simbolo €  (es. "123,45 €" o "€ 123,45")
      2. Numero decimale generico (es. "123,45" o "123.45")
    Strategia 3:
      text_only=False → primo token non vuoto dopo la chiave (testo normalizzato)
      text_only=True  → resto della riga dopo la chiave (testo grezzo, preserva spazi interni)
    """
    if not text:
        return ""

    escaped_key = re.escape(key)

    if text_only:
        # Lavora sul testo grezzo sfruttando due delimitatori naturali:
        #   - \n   → fine riga (PDF a singola colonna, HTML con tag blocco da _HTMLStripper)
        #   - \s{2,} → spazi multipli (separatore di colonna nei PDF a doppia colonna)
        # \s*[:\-]?\s* assorbe il separatore tra chiave e valore (spazi, ":", "–", \n).
        m = re.search(
            rf"{escaped_key}\s*[:\-]?\s*(.+?)(?=\s{{2,}}|[\r\n]|$)",
            text, re.IGNORECASE | re.MULTILINE
        )
        if m:
            val = m.group(1).strip()
            if val:
                return val
        # Fallback: fino a 6 token su testo normalizzato (HTML piatto senza \n né doppi spazi).
        normalized_fb = re.sub(r"\s+", " ", text)
        m = re.search(
            rf"{escaped_key}\s*[:\-]?\s*(\S+(?:\s+\S+){{0,5}})",
            normalized_fb, re.IGNORECASE
        )
        return m.group(1).strip() if m else ""

    normalized = re.sub(r"\s+", " ", text)

    # 1. importo con €
    for pattern in (
        rf"{escaped_key}[^0-9€]*?([\d]{{1,6}}[.,][\d]{{2}})\s*€",
        rf"{escaped_key}[^€]*?€\s*([\d]{{1,6}}[.,][\d]{{2}})",
    ):
        m = re.search(pattern, normalized, re.IGNORECASE)
        if m:
            return f"€ {m.group(1).strip()}"

    # 2. numero decimale senza €
    m = re.search(
        rf"{escaped_key}[^0-9]*?([\d]{{1,6}}[.,][\d]{{2}})",
        normalized, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()

    # 3. primo token non vuoto dopo la chiave
    m = re.search(
        rf"{escaped_key}\s*[:\-]?\s*(\S{{1,60}})",
        normalized, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()

    return ""


# ---------------------------------------------------------------------------
# Estrazione testo da allegati PDF
# ---------------------------------------------------------------------------

def _iter_parts(payload):
    """Itera ricorsivamente su tutte le parti foglia del payload Gmail."""
    parts = payload.get("parts", [])
    if not parts:
        yield payload
    else:
        for part in parts:
            yield from _iter_parts(part)


def _pdf_bytes_to_text(pdf_bytes):
    """Estrae il testo da un PDF in memoria usando pdfplumber.
    layout=True preserva le distanze reali tra caratteri: nei PDF a più colonne
    produce spazi multipli tra le colonne, utili come delimitatori in text_only.
    """
    import pdfplumber
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            pages = []
            for page in pdf.pages:
                try:
                    text = page.extract_text(layout=True) or ""
                except TypeError:
                    text = page.extract_text() or ""
                pages.append(text)
            return "\n".join(pages)
    except Exception as e:
        print(f"    Errore lettura PDF: {e}")
        return ""


def _find_pdf_parts(payload):
    """Restituisce lista di (part, filename) per tutti gli allegati PDF nel messaggio."""
    results = []
    for part in _iter_parts(payload):
        filename  = part.get("filename", "")
        mime_type = part.get("mimeType", "")
        is_pdf    = mime_type == "application/pdf" or filename.lower().endswith(".pdf")
        body      = part.get("body", {})
        has_data  = bool(body.get("data") or body.get("attachmentId"))
        if is_pdf and has_data:
            results.append((part, filename or "allegato.pdf"))
    return results


def _download_pdf_text(service, msg_id, part):
    """Scarica un singolo allegato PDF e ne restituisce il testo."""
    body          = part.get("body", {})
    inline_data   = body.get("data", "")
    attachment_id = body.get("attachmentId", "")
    if inline_data:
        pdf_bytes = base64.urlsafe_b64decode(inline_data + "==")
    elif attachment_id:
        att = service.users().messages().attachments().get(
            userId="me", messageId=msg_id, id=attachment_id
        ).execute()
        pdf_bytes = base64.urlsafe_b64decode(att["data"] + "==")
    else:
        return ""
    return _pdf_bytes_to_text(pdf_bytes)


# ---------------------------------------------------------------------------
# Estrazione con log
# ---------------------------------------------------------------------------

def _extract_with_log(service, msg_id, payload, keys, extract_from, link_text=None, text_only_keys=None):
    """
    Estrae i valori delle chiavi e genera un log descrittivo per ogni chiave.

    extract_from:
      "body"  → cerca solo nel corpo email
      "pdf"   → cerca solo negli allegati PDF
      "both"  → cerca nel corpo + PDF (testo unificato)
      "link"  → segue il link/pulsante nel corpo HTML e scarica il documento collegato
      None/"" → automatico: corpo prima, poi PDF se la chiave non è trovata

    link_text:      testo (o sottostringa) del pulsante da seguire (usato con "link").
    text_only_keys: set di chiavi che devono saltare le strategie monetarie (€/decimale).

    Restituisce (extracted: dict, search_log: str).
    """
    if not keys:
        return {}, ""

    pdf_parts = _find_pdf_parts(payload)
    has_pdfs  = bool(pdf_parts)
    pdf_names = ", ".join(fname for _, fname in pdf_parts) if has_pdfs else "—"

    extracted = {}
    log_parts = []

    # Lazy loaders — ogni sorgente viene calcolata/scaricata solo al primo accesso
    _cache = {"body": None, "pdf": None, "link_url": None, "link_text": None}

    def _body():
        if _cache["body"] is None:
            _cache["body"] = extract_body_text(payload)
        return _cache["body"]

    def _pdf():
        if _cache["pdf"] is None:
            if has_pdfs:
                texts = [_download_pdf_text(service, msg_id, p) for p, _ in pdf_parts]
                _cache["pdf"] = "\n".join(t for t in texts if t)
            else:
                _cache["pdf"] = ""
        return _cache["pdf"]

    def _link_url():
        if _cache["link_url"] is None:
            links = _find_links(payload)
            _cache["link_url"] = _select_link(links, link_text) or ""
            if not _cache["link_url"]:
                _cache["link_url"] = ""   # evita None
            _cache["_link_count"] = len(links)
        return _cache["link_url"]

    def _link_content():
        if _cache["link_text"] is None:
            url = _link_url()
            _cache["link_text"] = _download_url_as_text(url) if url else ""
        return _cache["link_text"]

    _text_only_keys = text_only_keys or set()

    for key in keys:
        tonly = key in _text_only_keys

        if extract_from == "body":
            val = extract_value_for_key(_body(), key, text_only=tonly)
            extracted[key] = val
            log_parts.append(f"{key}: corpo {'✓' if val else '✗ non trovato'}")

        elif extract_from == "pdf":
            val = extract_value_for_key(_pdf(), key, text_only=tonly)
            extracted[key] = val
            if val:
                log_parts.append(f"{key}: PDF ✓ ({pdf_names})")
            elif has_pdfs:
                log_parts.append(f"{key}: PDF ✗ non trovato ({pdf_names})")
            else:
                log_parts.append(f"{key}: PDF ✗ nessun allegato")

        elif extract_from == "both":
            val = extract_value_for_key(_body() + "\n" + _pdf(), key, text_only=tonly)
            extracted[key] = val
            src = "corpo+PDF" if has_pdfs else "corpo"
            log_parts.append(f"{key}: {src} {'✓' if val else '✗ non trovato'}")

        elif extract_from == "link":
            url = _link_url()
            val = extract_value_for_key(_link_content(), key, text_only=tonly)
            extracted[key] = val
            short_url = (url[:60] + "…") if len(url) > 60 else url
            if val:
                log_parts.append(f"{key}: link ✓ ({short_url})")
            elif url:
                log_parts.append(f"{key}: link ✗ non trovato ({short_url})")
            else:
                n = _cache.get("_link_count", 0)
                log_parts.append(
                    f"{key}: link ✗ nessun link corrispondente "
                    f"({'link_text='+repr(link_text) if link_text else str(n)+' link trovati'})"
                )

        else:  # auto
            val = extract_value_for_key(_body(), key, text_only=tonly)
            if val:
                extracted[key] = val
                log_parts.append(f"{key}: corpo ✓")
            else:
                val = extract_value_for_key(_pdf(), key, text_only=tonly)
                if val:
                    extracted[key] = val
                    log_parts.append(f"{key}: PDF ✓ ({pdf_names})")
                elif has_pdfs:
                    extracted[key] = ""
                    log_parts.append(f"{key}: ✗ non trovato (PDF: {pdf_names})")
                else:
                    extracted[key] = ""
                    log_parts.append(f"{key}: ✗ non trovato (no allegati PDF)")

    return extracted, " | ".join(log_parts)


# ---------------------------------------------------------------------------
# Recupero email
# ---------------------------------------------------------------------------

def _build_query(sender, date_range):
    """Costruisce la query Gmail con filtro mittente e range di date."""
    q = f"from:{sender}"
    if date_range:
        if date_range.get("from"):
            q += f" after:{date_range['from'].strftime('%Y/%m/%d')}"
        if date_range.get("to"):
            # Gmail before: è esclusivo — aggiungiamo 1 giorno per includere la data finale
            from datetime import timedelta
            to_inclusive = date_range["to"] + timedelta(days=1)
            q += f" before:{to_inclusive.strftime('%Y/%m/%d')}"
    return q


def _list_message_ids(service, sender, date_range=None):
    ids, page_token = [], None
    while True:
        params = {"userId": "me", "maxResults": BATCH_SIZE, "q": _build_query(sender, date_range)}
        if page_token:
            params["pageToken"] = page_token
        results = service.users().messages().list(**params).execute()
        ids.extend(results.get("messages", []))
        page_token = results.get("nextPageToken")
        if not page_token:
            break
    return ids


def fetch_all_emails(service, senders_config, date_range=None):
    """Recupera email per ogni mittente configurato ed estrae le chiavi richieste."""
    all_emails = []

    for sender_cfg in senders_config:
        sender       = sender_cfg["email"]
        keys          = sender_cfg.get("extract_bill", [])
        extract_from  = sender_cfg.get("extract_from", "")  # "" = auto
        link_text     = sender_cfg.get("link_text", None)
        supply_labels = sender_cfg.get("supply_labels", [])
        customer_key  = sender_cfg.get("extract_customer", "")
        period_key    = sender_cfg.get("extract_period", "")
        needs_full    = bool(keys) or bool(supply_labels) or bool(customer_key) or bool(period_key)

        src_label = extract_from if extract_from else "auto"
        print(f"\nRicerca email da: {sender}")
        if keys:
            extra = f"  link_text={repr(link_text)}" if extract_from == "link" and link_text else ""
            print(f"  Chiavi da estrarre: {', '.join(keys)}  [sorgente: {src_label}{extra}]")

        msg_ids = _list_message_ids(service, sender, date_range)
        print(f"  Trovati {len(msg_ids)} messaggi — scaricamento dettagli...")

        for i, msg_info in enumerate(msg_ids, 1):
            try:
                fmt    = "full" if needs_full else "metadata"
                kwargs = {"userId": "me", "id": msg_info["id"], "format": fmt}
                if not needs_full:
                    kwargs["metadataHeaders"] = ["From", "Subject", "Date"]

                msg     = service.users().messages().get(**kwargs).execute()
                payload = msg.get("payload", {})
                headers = {
                    h["name"].lower(): h["value"]
                    for h in payload.get("headers", [])
                }

                extra_keys = [k for k in (customer_key, period_key) if k and k not in keys]
                all_keys   = keys + extra_keys
                text_only  = {k for k in (customer_key, period_key) if k}
                extracted, search_log = _extract_with_log(
                    service, msg_info["id"], payload, all_keys, extract_from, link_text,
                    text_only_keys=text_only or None,
                )
                customer_value = extracted.pop(customer_key, "") if customer_key else ""
                period_value   = extracted.pop(period_key,   "") if period_key   else ""

                supply_label = ""
                if supply_labels:
                    subject_text = headers.get("subject", "")
                    body_text    = extract_body_text(payload)
                    search_text  = (subject_text + " " + body_text).upper()
                    for label in supply_labels:
                        if label.upper() in search_text:
                            supply_label = label
                            break

                all_emails.append({
                    "sender_email":   sender,
                    "from":           headers.get("from", sender),
                    "subject":        headers.get("subject", "(nessun oggetto)"),
                    "date_raw":       headers.get("date", ""),
                    "date":           parse_date(headers.get("date", "")),
                    "snippet":        msg.get("snippet", ""),
                    "extracted":      extracted,
                    "search_log":     search_log,
                    "supply_label":   supply_label,
                    "customer_value": customer_value,
                    "period_value":   period_value,
                })

                if i % 50 == 0 or i == len(msg_ids):
                    print(f"  {i}/{len(msg_ids)} elaborati")

            except Exception as e:
                print(f"  Errore nel messaggio {msg_info['id']}: {e}")

    all_emails.sort(key=lambda x: x["date"] or datetime.min, reverse=True)
    print(f"\nRecupero completato: {len(all_emails)} email totali.\n")
    return all_emails


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _make_sheet_name(sender_email, existing_names):
    """
    Genera un nome foglio valido per Excel (max 31 chars, no \\ / ? * [ ] :).
    Se il nome è già usato aggiunge un suffisso numerico.
    """
    clean = re.sub(r'[\\/?*\[\]:]', '_', sender_email)[:31]
    if clean not in existing_names:
        return clean
    for i in range(2, 100):
        candidate = f"{clean[:28]}_{i}"
        if candidate not in existing_names:
            return candidate
    return clean  # fallback (non dovrebbe mai accadere)


def _write_sheet(wb, sheet_name, sender_emails, keys, styles, has_supply_label=False, customer_key="", period_key=""):
    """Crea un foglio Excel per un singolo mittente."""
    from openpyxl.utils import get_column_letter

    header_font, header_fill, header_align, \
    data_font, value_font, log_font, \
    center_top, left_wrap, right_top, \
    thin_border, alt_fill, log_alt_fill = styles

    ws = wb.create_sheet(title=sheet_name)

    # Colonne: senza "Mittente" (ridondante per foglio dedicato)
    fixed_cols = [
        ("#",          5,  center_top, data_font),
        ("Data",      22,  center_top, data_font),
        ("Oggetto",   55,  left_wrap,  data_font),
        ("Anteprima", 75,  left_wrap,  data_font),
    ]
    supply_col    = [("Tipo fornitura", 20, center_top, data_font)] if has_supply_label else []
    customer_col  = [("Cliente",        30, left_wrap,  data_font)] if customer_key      else []
    period_col    = [("Periodo",        22, center_top, data_font)] if period_key         else []
    dynamic_cols  = [(key, 22, right_top, value_font) for key in keys]
    totale_col    = [("TOTALE DEFINITIVO", 22, right_top, value_font)]
    log_col       = [("Log ricerca", 55, left_wrap, log_font)]
    all_cols      = fixed_cols + supply_col + customer_col + period_col + dynamic_cols + totale_col + log_col

    for col, (title, width, *_) in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20

    for idx, email in enumerate(sender_emails, 1):
        row      = idx + 1
        fill     = alt_fill     if idx % 2 == 0 else None
        log_fill = log_alt_fill if idx % 2 == 0 else None

        date_val = email["date"]
        date_str = date_val.strftime("%d/%m/%Y %H:%M") if date_val else email["date_raw"]

        row_values = [
            (idx,               center_top, data_font,  fill),
            (date_str,          center_top, data_font,  fill),
            (email["subject"],  left_wrap,  data_font,  fill),
            (email["snippet"],  left_wrap,  data_font,  fill),
        ]
        if has_supply_label:
            row_values.append((email.get("supply_label", ""), center_top, data_font, fill))
        if customer_key:
            row_values.append((email.get("customer_value", ""), left_wrap,  data_font, fill))
        if period_key:
            row_values.append((email.get("period_value",   ""), center_top, data_font, fill))
        for key in keys:
            row_values.append((email["extracted"].get(key, ""), right_top, value_font, fill))
        if keys:
            amounts   = [_parse_amount(email["extracted"].get(k, "")) for k in keys]
            valid     = [a for a in amounts if a is not None]
            max_val   = max(valid) if valid else None
            if max_val is not None:
                # ripristina il formato del valore originale (con € se presente)
                orig = next(
                    email["extracted"].get(k, "") for k in keys
                    if _parse_amount(email["extracted"].get(k, "")) == max_val
                )
                totale_def = orig
            else:
                totale_def = ""
            row_values.append((totale_def, right_top, value_font, fill))
        row_values.append((email.get("search_log", ""), left_wrap, log_font, log_fill))

        for col, (value, align, font, cell_fill) in enumerate(row_values, 1):
            c            = ws.cell(row=row, column=col, value=value)
            c.font       = font
            c.alignment  = align
            c.border     = thin_border
            if cell_fill:
                c.fill = cell_fill

        ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(all_cols))
    ws.auto_filter.ref = f"A1:{last_col}{len(sender_emails) + 1}"


def create_excel_report(emails, senders_config, output_file):
    """
    Crea il file Excel con un foglio per ogni mittente.
    Colonne per foglio: #, Data, Oggetto, Anteprima, [chiavi mittente], Log ricerca
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict

    # Stili condivisi tra tutti i fogli
    styles = (
        Font(name="Arial", bold=True, color="FFFFFF", size=11),   # header_font
        PatternFill("solid", fgColor="C0392B"),                    # header_fill
        Alignment(horizontal="center", vertical="center", wrap_text=True),  # header_align
        Font(name="Arial", size=10),                               # data_font
        Font(name="Arial", size=10, bold=True, color="1A5276"),    # value_font
        Font(name="Arial", size=9, italic=True, color="555555"),   # log_font
        Alignment(horizontal="center", vertical="top"),            # center_top
        Alignment(horizontal="left",   vertical="top", wrap_text=True),     # left_wrap
        Alignment(horizontal="right",  vertical="top"),            # right_top
        Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        ),                                                          # thin_border
        PatternFill("solid", fgColor="FDECEA"),                    # alt_fill
        PatternFill("solid", fgColor="F5F5F5"),                    # log_alt_fill
    )

    # Raggruppa email per mittente mantenendo l'ordine di senders_config
    emails_by_sender = defaultdict(list)
    for email in emails:
        emails_by_sender[email["sender_email"]].append(email)

    keys_by_sender = {s["email"]: s.get("extract_bill", []) for s in senders_config}

    wb = Workbook()
    wb.remove(wb.active)   # rimuove il foglio vuoto di default

    sheet_names = []
    for sender_cfg in senders_config:
        sender         = sender_cfg["email"]
        keys           = keys_by_sender.get(sender, [])
        sender_emails  = emails_by_sender.get(sender, [])
        sheet_name     = _make_sheet_name(sender, sheet_names)
        sheet_names.append(sheet_name)

        has_supply   = bool(sender_cfg.get("supply_labels"))
        customer_key = sender_cfg.get("extract_customer", "")
        period_key   = sender_cfg.get("extract_period",   "")
        _write_sheet(wb, sheet_name, sender_emails, keys, styles, has_supply_label=has_supply, customer_key=customer_key, period_key=period_key)
        print(f"  Foglio '{sheet_name}': {len(sender_emails)} email")

    wb.save(output_file)
    print(f"\nFile Excel salvato: {output_file}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    senders_config, date_range = load_settings()
    timestamp                  = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file                = f"enel_emails_{timestamp}.xlsx"

    print("=" * 60)
    print("   Gmail Utilities Bill Retrieve")
    for s in senders_config:
        keys_str = ", ".join(s.get("extract_bill", [])) or "—"
        src      = s.get("extract_from") or "auto"
        print(f"   • {s['email']}  [{keys_str}]  (sorgente: {src})")
    dr_from = date_range["from"].strftime("%d/%m/%Y") if date_range["from"] else "—"
    dr_to   = date_range["to"].strftime("%d/%m/%Y")   if date_range["to"]   else "—"
    if date_range["from"] or date_range["to"]:
        print(f"   Range date: dal {dr_from} al {dr_to}")
    print("=" * 60)

    creds   = authenticate()
    service = build("gmail", "v1", credentials=creds)

    profile = service.users().getProfile(userId="me").execute()
    print(f"\nAccount: {profile.get('emailAddress', 'N/A')}")

    emails = fetch_all_emails(service, senders_config, date_range)

    if not emails:
        print("Nessuna email trovata.")
        return

    print("Generazione fogli Excel:")
    create_excel_report(emails, senders_config, output_file)

    print("\nPrime 10 email (più recenti):")
    print("-" * 60)
    for i, email in enumerate(emails[:10], 1):
        date_str  = email["date"].strftime("%d/%m/%Y") if email["date"] else "N/D"
        extras    = "  →  " + "  |  ".join(
            f"{k}: {v}" for k, v in email["extracted"].items() if v
        ) if email["extracted"] else ""
        print(f"  {i:2}. [{date_str}] {email['subject'][:40]}{extras}")

    print(f"\nFile generato: {output_file}")


if __name__ == "__main__":
    main()
